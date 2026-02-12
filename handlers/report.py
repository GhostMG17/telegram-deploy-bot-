import tempfile
from telegram import Update, InputFile
from telegram.ext import ContextTypes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from database import db as db

async def report(update: Update, context: ContextTypes.DEFAULT_TYPE, query=None):
    data, tasks = db.get_report_table()
    if not data:
        text = "Пока нет данных для отчёта."
        if query:
            await query.message.reply_text(text)
        else:
            await update.message.reply_text(text)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Ramadan Progress"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # синий
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = ["Имя пользователя", "Дата"] + tasks + ["Все задачи выполнены?"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 20  # ширина колонок

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(headers, start=1):
            value = row_data.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = border

            if col_name == "Все задачи выполнены?":
                if value == "Да":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")  # светло-зелёный
                elif value == "Нет":
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")  # светло-красный

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    with open(tmp_path, "rb") as f:
        if query:
            await query.message.reply_document(InputFile(f, filename="Ramadan_Report.xlsx"))
        else:
            await update.message.reply_document(InputFile(f, filename="Ramadan_Report.xlsx"))